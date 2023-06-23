import re
import openpyxl

# 读取Excel文件
wb = openpyxl.load_workbook('数据.xlsx')

# 读取985大学和211大学的名单
with open('985.txt', 'r', encoding='utf-8') as f:
    universities_985 = [line.strip() for line in f.readlines()]

with open('211.txt', 'r', encoding='utf-8') as f:
    universities_211 = [line.strip() for line in f.readlines()]

# 遍历每个表
for sheet_name in wb.sheetnames:
    # 选择当前表
    ws = wb[sheet_name]

    # 遍历每一行，对高校名称进行处理
    for row in ws.iter_rows(min_row=2):
        school_name = row[1].value
        if school_name:
            if school_name in universities_985:
                row[1].value = school_name + '985'
            elif school_name in universities_211:
                row[1].value = school_name + '211'

# 保存修改后的Excel文件
wb.save('数据.xlsx')
