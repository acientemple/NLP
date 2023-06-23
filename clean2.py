import re
import openpyxl

# 读取Excel文件
wb = openpyxl.load_workbook('数据.xlsx')

# 选择工作表
# ws = wb.active
for sheet_name in wb.sheetnames:
    # 选择当前表
    ws = wb[sheet_name]
# 遍历每一行，去掉学校名称中大学或者学院后面的内容
    for row in ws.iter_rows(min_row=2):
        school_name = row[3].value
        if school_name:
            new_school_name = re.sub(r'（.*', '', school_name)
            row[3].value = new_school_name

# 保存修改后的Excel文件
wb.save('数据.xlsx')
